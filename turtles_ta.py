import os
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
TICKERS_FILE = os.path.join(BASE_DIR, "tickers.csv")
OUTPUT_FILE  = os.path.join(BASE_DIR, "historical_data.xlsx")

# ── Date range: last 5 years ──────────────────────────────────────────────────
end_date   = datetime.today()
start_date = end_date - timedelta(days=5 * 365)

# ── Tickers: SPY always first ─────────────────────────────────────────────────
tickers_df  = pd.read_csv(TICKERS_FILE, header=None, names=["ticker"])
csv_tickers = tickers_df["ticker"].str.strip().tolist()
tickers     = ["SPY"] + [t for t in csv_tickers if t != "SPY"]
print(f"Tickers: {tickers}")

# ── Styles ────────────────────────────────────────────────────────────────────
def solid_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

FILL_BLUE       = solid_fill("4472C4")
FILL_GREEN      = solid_fill("70AD47")
FILL_ORANGE     = solid_fill("ED7D31")
FILL_YELLOW     = solid_fill("FFD966")
FILL_GREEN_CELL = solid_fill("C6EFCE")
FILL_RED_CELL   = solid_fill("FFC7CE")
FONT_WHITE      = Font(color="FFFFFF", bold=True)
FONT_WHITE_REG  = Font(color="FFFFFF")

# ── Column layout (data sheets) ───────────────────────────────────────────────
# A=Date  B=Open  C=High  D=Low  E=Close  F=Volume
# G=Long Range   H=Short Range
# I=Long N       J=Short N
# K=20 Day Low   L=20 Day High
# M=20 Day Low (High) 120D
# N=20 Day Return %   O=250 Day Return %
# P= (gap)
# Q=label  R=value  S=dollar Δ  T=note  (summary panel)
# U= (gap)
# V=label  W=value  (right info panel)

DATA_HEADERS = {
    "G": "Long Range",
    "H": "Short Range",
    "I": "Long N",
    "J": "Short N",
    "K": "20 Day Low",
    "L": "20 Day High",
    "M": "20 Day Low (High) 120D",
    "N": "20 Day Return",
    "O": "250 Day Return",
}

# ── Row formula builder ───────────────────────────────────────────────────────
def cap(row, offset, last_row):
    return min(row + offset, last_row)

def build_row_formulas(r, last_row):
    c20  = cap(r, 19,  last_row)
    c120 = cap(r, 119, last_row)
    c249 = cap(r, 249, last_row)
    l40  = cap(r, 39,  last_row)
    l20  = cap(r, 19,  last_row)
    return {
        "G": f"=MAX(C{r}:C{c20})-MIN(D{r}:D{l40})",
        "H": f"=MAX(C{r}:C{c120})-MIN(D{r}:D{l20})",
        "I": f"=G{r}/5",
        "J": f"=H{r}/5",
        "K": f"=MIN(D{r}:D{l20})",
        "L": f"=MAX(C{r}:C{c20})",
        "M": f"=MAX(K{r}:K{c120})",
        "N": f"=IFERROR((E{r}-E{c20})/E{c20},\"\")",
        "O": f"=IFERROR((E{r}-E{c249})/E{c249},\"\")",
    }

# ── Summary panel writer ──────────────────────────────────────────────────────
def write_summary_panel(ws, ticker, is_spy=False):
    ws["Q1"] = ticker
    ws["Q1"].font = Font(bold=True, size=13)

    daily_rows = {
        3: ("Daily High",    "=C2", "$#,##0.00"),
        4: ("Current Price", "=E2", "$#,##0.00"),
        5: ("Daily Low",     "=D2", "$#,##0.00"),
        6: ("Volume",        "=F2", "#,##0"),
    }
    for row, (label, formula, fmt) in daily_rows.items():
        ws[f"Q{row}"].value = label
        ws[f"Q{row}"].fill  = FILL_BLUE
        ws[f"Q{row}"].font  = FONT_WHITE
        ws[f"R{row}"].value = formula
        ws[f"R{row}"].number_format = fmt

    ws["Q8"]  = "Long Risk $$"
    ws["Q9"]  = "Long Risk $$/N"
    ws["Q10"] = "Short Risk $$"
    ws["Q11"] = "Short Risk $$/N"

    ws["R8"].value  = "=IFERROR(E22-E252,\"\")"
    ws["R9"].value  = "=IFERROR(R8/(I2*Settings!$B$2),\"\")"
    ws["R10"].value = "=IF(R4-MAX(C2:C121)>0,0,R4-MAX(C2:C121))"
    ws["R11"].value = "=IFERROR(R10/(J2+Settings!$C$2),\"\")"
    for row in (8, 9, 10, 11):
        ws[f"R{row}"].number_format = "$#,##0.00"

    # 20D Residual Return — S13 = E22−E2 (negative = price rose = Inverse)
    ws["Q13"].value = '="20D Residual Return ("&IF(S13<0,"Inverse","Positive")&")"'
    ws["Q13"].fill  = FILL_GREEN
    ws["Q13"].font  = FONT_WHITE_REG
    ws["R13"].value = "=IFERROR(E22,\"\")"
    ws["R13"].number_format = "$#,##0.00"
    ws["S13"].value = "=IFERROR(E22-E2,\"\")"
    ws["S13"].number_format = "$#,##0.00"
    ws["T13"].value = "20 Price"

    # 250D Residual Return — S14 = E2−E252 (positive = price rose = Positive)
    ws["Q14"].value = '="250D Residual Return ("&IF(S14>0,"Positive","Inverse")&")"'
    ws["Q14"].fill  = FILL_ORANGE
    ws["Q14"].font  = FONT_WHITE_REG
    ws["R14"].value = "=IFERROR(E252,\"\")"
    ws["R14"].number_format = "$#,##0.00"
    ws["S14"].value = "=IFERROR(E2-E252,\"\")"
    ws["S14"].number_format = "$#,##0.00"
    ws["T14"].value = "250 Price"

    ws["Q16"] = "20 Day Low (High) (120D)"
    ws["R16"].value = "=M2"
    ws["R16"].number_format = "$#,##0.00"

    ws["Q17"] = "Points to 20 Day Low / Long ATR"
    ws["R17"].value = "=IFERROR((R4-K2)/I2,\"\")"
    ws["R17"].number_format = "0.00"

    ws["Q18"] = "% Price from Highest 20 Low"
    ws["R18"].value = "=IFERROR((R4-M2)/M2,\"\")"
    ws["R18"].number_format = "0.00%"

    if is_spy:
        ws["Q20"] = "20 Day Return"
        ws["R20"].value = "=N2"
        ws["R20"].number_format = "0.00%"
        ws["Q21"] = "250 Day Return"
        ws["R21"].value = "=O2"
        ws["R21"].number_format = "0.00%"

    # Right panel
    ws["V2"] = "Stock Details"
    ws["V2"].font = Font(bold=True)

    ws["V3"] = "Dollars Per Point"
    ws["W3"].value = "=Settings!$D$2"
    ws["W3"].number_format = "0.00"

    ws["V4"] = "Long Range: 20/40"
    ws["V4"].fill = FILL_GREEN
    ws["V4"].font = FONT_WHITE
    ws["V5"] = "Long N"
    ws["W5"].value = "=I2"
    ws["W5"].number_format = "0.00"
    ws["V6"] = "Long NS$"
    ws["W6"].value = "=IFERROR(I2*Settings!$D$2,\"\")"
    ws["W6"].number_format = "$#,##0.00"

    ws["V7"] = "Short Range"
    ws["V7"].fill = FILL_ORANGE
    ws["V7"].font = FONT_WHITE
    ws["V8"] = "Short N"
    ws["W8"].value = "=J2"
    ws["W8"].number_format = "0.00"
    ws["V9"] = "Short NS$"
    ws["W9"].value = "=IFERROR(J2*Settings!$D$2,\"\")"
    ws["W9"].number_format = "$#,##0.00"

    ws.column_dimensions["Q"].width = 38
    ws.column_dimensions["R"].width = 16
    ws.column_dimensions["S"].width = 14
    ws.column_dimensions["T"].width = 12
    ws.column_dimensions["V"].width = 22
    ws.column_dimensions["W"].width = 14


# ── Settings sheet builder ────────────────────────────────────────────────────
def build_settings_sheet(wb, sectors=None, industries=None):
    """
    All formula-referenced values live in row 2, varying columns:
      $B$2 = Long Risk Multiplier      (default 8)
      $C$2 = Short Risk Additive       (default 0.5)
      $D$2 = Dollars Per Point         (default 1)
      $E$2 = Market Cap threshold      (default 10 000 000 000)
      $F$2 = PE Ratio alert threshold  (default 30)
      $G$2 = Sector filter             (blank = all)
      $H$2 = Industry filter           (blank = all)

    Dropdown lists for sector/industry are written in cols J and K.
    """
    ws = wb["Settings"] if "Settings" in wb.sheetnames else wb.create_sheet("Settings")

    ws["A1"] = "Settings — change the yellow cells to update all sheets live"
    ws["A1"].font = Font(bold=True, size=13)

    config = [
        # (row, label_col_A, value_col, value, fmt, note)
        (2,  "Long Risk Multiplier  →  Long Risk $$/N denominator",       "B", 8,              "0.00",        ""),
        (3,  "Short Risk Additive   →  Short Risk $$/N denominator offset","C", 0.5,            "0.00",        "shares row 2"),
        (4,  "Dollars Per Point     →  NS$ multiplier",                    "D", 1,              "0.00",        "shares row 2"),
        (5,  "Market Cap Filter     →  Mrk. Cap < this value",             "E", 10_000_000_000, "$#,##0",      "shares row 2"),
        (6,  "PE Ratio Alert        →  highlight PE above this threshold", "F", 30,             "0.00",        "shares row 2"),
        (7,  "Sector Filter         →  leave blank to show all sectors",   "G", "",             "@",           "dropdown ↓"),
        (8,  "Industry Filter       →  leave blank to show all industries","H", "",             "@",           "dropdown ↓"),
    ]

    for row, label, val_col, value, fmt, note in config:
        ws[f"A{row}"] = label
        cell = ws[f"{val_col}2"]   # all values in row 2
        cell.value = value
        cell.number_format = fmt
        cell.fill = FILL_YELLOW
        if note:
            ws[f"I{row}"] = note

    ws.column_dimensions["A"].width = 58
    for col in ("B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 18

    # ── Sector / Industry dropdown lists (cols J and K) ───────────────
    if sectors:
        ws["J1"] = "Sectors (auto-populated)"
        ws["J1"].font = Font(bold=True, italic=True, color="888888")
        for i, s in enumerate(sectors, 2):
            ws.cell(row=i, column=10, value=s)
        last_sector_row = len(sectors) + 1
        dv_sector = DataValidation(
            type="list",
            formula1=f"Settings!$J$2:$J${last_sector_row}",
            allow_blank=True,
            showDropDown=False,
        )
        dv_sector.sqref = "G2"
        ws.add_data_validation(dv_sector)

    if industries:
        ws["K1"] = "Industries (auto-populated)"
        ws["K1"].font = Font(bold=True, italic=True, color="888888")
        for i, ind in enumerate(industries, 2):
            ws.cell(row=i, column=11, value=ind)
        last_industry_row = len(industries) + 1
        dv_industry = DataValidation(
            type="list",
            formula1=f"Settings!$K$2:$K${last_industry_row}",
            allow_blank=True,
            showDropDown=False,
        )
        dv_industry.sqref = "H2"
        ws.add_data_validation(dv_industry)

    ws.column_dimensions["J"].width = 30
    ws.column_dimensions["K"].width = 35


# ── Dashboard builder ─────────────────────────────────────────────────────────
def build_dashboard(wb, tickers, ticker_meta):
    """
    One row per ticker.

    Columns
    -------
    A  Stock (name + ticker)          static
    B  Sector                         static
    C  Industry                       static
    D  Current Price                  live  → '{ticker}'!R4
    E  Market Cap                     static
    F  Mrk. Cap < (Yes/No)            formula vs Settings!$E$2
    G  PE Ratio                       static
    H  PE > Threshold (Yes/No)        formula vs Settings!$F$2
    I  Beta                           static
    J  Short Trend Index              live  → '{ticker}'!R11
    K  CCM Trend Index Long           live  → '{ticker}'!R9
    L  CCM TREND INDEX                formula = J + K  (green/red highlight)
    M  Avg 20D Volume                 live  → AVERAGE('{ticker}'!F2:F21)
    N  52W High                       live  → MAX('{ticker}'!C2:C253)
    O  52W Low                        live  → MIN('{ticker}'!D2:D253)
    P  Passes All Filters             formula (✓ / blank)
    """
    ws = wb.create_sheet("Dashboard")

    headers = [
        "Stock", "Sector", "Industry",
        "Current Price", "Market Cap", "Mrk. Cap <",
        "PE Ratio", "PE > Threshold",
        "Beta",
        "Short Trend Index", "CCM Trend Index Long", "CCM TREND INDEX",
        "Avg 20D Volume", "52W High", "52W Low",
        "Passes All Filters",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = FONT_WHITE
        cell.fill  = FILL_BLUE

    col_widths = [48, 22, 30, 14, 22, 12, 10, 14, 8, 18, 22, 18, 18, 12, 12, 18]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, ticker in enumerate(tickers, 2):
        meta = ticker_meta.get(ticker, {})
        t    = ticker.replace("'", "''")   # escape any apostrophes in sheet name

        # A: Stock name
        ws.cell(row=row_idx, column=1,
                value=f"{meta.get('name', ticker)} ({ticker})")

        # B: Sector, C: Industry
        ws.cell(row=row_idx, column=2, value=meta.get("sector",   "N/A"))
        ws.cell(row=row_idx, column=3, value=meta.get("industry", "N/A"))

        # D: Current Price (live)
        c = ws.cell(row=row_idx, column=4, value=f"='{t}'!R4")
        c.number_format = "$#,##0.00"

        # E: Market Cap (static)
        mc = meta.get("market_cap")
        c  = ws.cell(row=row_idx, column=5, value=mc)
        c.number_format = '$#,##0.00'

        # F: Mrk. Cap < filter
        ws.cell(row=row_idx, column=6,
                value=f'=IF(E{row_idx}="","N/A",IF(E{row_idx}<Settings!$E$2,"Yes","No"))')

        # G: PE Ratio (static)
        pe = meta.get("pe_ratio")
        c  = ws.cell(row=row_idx, column=7, value=pe)
        if pe is not None:
            c.number_format = "0.00"

        # H: PE > Threshold
        ws.cell(row=row_idx, column=8,
                value=f'=IF(G{row_idx}="","N/A",IF(G{row_idx}>Settings!$F$2,"Yes","No"))')

        # I: Beta (static)
        beta = meta.get("beta")
        c    = ws.cell(row=row_idx, column=9, value=beta)
        if beta is not None:
            c.number_format = "0.00"

        # J: Short Trend Index (live)
        c = ws.cell(row=row_idx, column=10, value=f"='{t}'!R11")
        c.number_format = "0.000"

        # K: CCM Trend Index Long (live)
        c = ws.cell(row=row_idx, column=11, value=f"='{t}'!R9")
        c.number_format = "0.000"

        # L: CCM TREND INDEX = J + K
        c = ws.cell(row=row_idx, column=12,
                    value=f"=IFERROR(J{row_idx}+K{row_idx},\"\")")
        c.number_format = "0.000"

        # M: Avg 20D Volume (live)
        c = ws.cell(row=row_idx, column=13,
                    value=f"=IFERROR(AVERAGE('{t}'!F2:F21),\"\")")
        c.number_format = "#,##0"

        # N: 52W High (live)
        c = ws.cell(row=row_idx, column=14,
                    value=f"=IFERROR(MAX('{t}'!C2:C253),\"\")")
        c.number_format = "$#,##0.00"

        # O: 52W Low (live)
        c = ws.cell(row=row_idx, column=15,
                    value=f"=IFERROR(MIN('{t}'!D2:D253),\"\")")
        c.number_format = "$#,##0.00"

        # P: Passes All Filters
        # Passes if: Market Cap filter met (or blank) AND Sector matches (or blank)
        #            AND Industry matches (or blank)
        ws.cell(row=row_idx, column=16,
                value=(
                    f'=IF(AND('
                    f'OR(Settings!$E$2="",E{row_idx}="",E{row_idx}<Settings!$E$2),'
                    f'OR(Settings!$G$2="",B{row_idx}=Settings!$G$2),'
                    f'OR(Settings!$H$2="",C{row_idx}=Settings!$H$2)'
                    f'),"✓","")'
                ))

    # ── Conditional formatting ────────────────────────────────────────
    last_data_row = len(tickers) + 1

    # CCM TREND INDEX (col L): green if > 0, red if < 0
    ws.conditional_formatting.add(
        f"L2:L{last_data_row}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=FILL_GREEN_CELL, font=Font(bold=True, color="276221")),
    )
    ws.conditional_formatting.add(
        f"L2:L{last_data_row}",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=FILL_RED_CELL, font=Font(bold=True, color="9C0006")),
    )

    # Mrk. Cap < (col F): green Yes, orange No
    ws.conditional_formatting.add(
        f"F2:F{last_data_row}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=FILL_GREEN_CELL),
    )
    ws.conditional_formatting.add(
        f"F2:F{last_data_row}",
        CellIsRule(operator="equal", formula=['"No"'], fill=FILL_ORANGE),
    )

    # PE > Threshold (col H): orange if Yes (high PE = caution)
    ws.conditional_formatting.add(
        f"H2:H{last_data_row}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=FILL_YELLOW),
    )

    # Passes All Filters (col P): green highlight when ✓
    ws.conditional_formatting.add(
        f"P2:P{last_data_row}",
        CellIsRule(operator="equal", formula=['"✓"'], fill=FILL_GREEN_CELL,
                   font=Font(bold=True, color="276221")),
    )

    # Freeze top row
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────
ticker_meta = {}   # {ticker: {name, sector, industry, market_cap, pe_ratio, beta}}

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    wb = writer.book

    # Remove openpyxl's default blank sheet
    for default_name in ("Sheet", "Sheet1"):
        if default_name in wb.sheetnames:
            del wb[default_name]

    # Create Settings shell now (values populated after meta is fetched)
    wb.create_sheet("Settings")

    for ticker in tickers:
        print(f"Fetching {ticker}...")
        try:
            # ── OHLCV ──────────────────────────────────────────────
            raw = yf.download(
                ticker,
                start=start_date.strftime("%Y-%m-%d"),
                end=end_date.strftime("%Y-%m-%d"),
                interval="1d",
                auto_adjust=True,
                progress=False,
            )
            if raw.empty:
                print(f"  WARNING: No data for {ticker}, skipping.")
                continue

            df = raw[["Open", "High", "Low", "Close", "Volume"]].copy()
            df.columns = ["Open", "High", "Low", "Close", "Volume"]
            df.index.name = "Date"
            df.reset_index(inplace=True)
            df["Date"] = pd.to_datetime(df["Date"]).dt.date
            df.sort_values("Date", ascending=False, inplace=True)
            df.reset_index(drop=True, inplace=True)

            df.to_excel(writer, sheet_name=ticker, index=False, float_format="%.4f")

            ws       = writer.sheets[ticker]
            n        = len(df)
            LAST_ROW = n + 1

            for col_letter, label in DATA_HEADERS.items():
                ws[f"{col_letter}1"] = label

            ws.column_dimensions["A"].width = 13
            for col in ("B", "C", "D", "E"):
                ws.column_dimensions[col].width = 11
            ws.column_dimensions["F"].width = 15
            for col in ("G", "H", "I", "J", "K", "L", "M", "N", "O"):
                ws.column_dimensions[col].width = 16

            for i in range(n):
                r        = i + 2
                formulas = build_row_formulas(r, LAST_ROW)
                for col_letter, formula in formulas.items():
                    cell = ws[f"{col_letter}{r}"]
                    cell.value = formula
                    if col_letter in ("N", "O"):
                        cell.number_format = "0.00%"

            write_summary_panel(ws, ticker, is_spy=(ticker == "SPY"))

            # ── Fundamental meta ───────────────────────────────────
            try:
                info = yf.Ticker(ticker).info
            except Exception:
                info = {}
            ticker_meta[ticker] = {
                "name":       info.get("longName") or info.get("shortName") or ticker,
                "sector":     info.get("sector",    "N/A"),
                "industry":   info.get("industry",  "N/A"),
                "market_cap": info.get("marketCap"),
                "pe_ratio":   info.get("trailingPE"),
                "beta":       info.get("beta"),
            }

            print(f"  {n} rows → sheet '{ticker}'  |  "
                  f"sector: {ticker_meta[ticker]['sector']}")

        except Exception as e:
            print(f"  ERROR fetching {ticker}: {e}")

    # ── Build Settings (now we have all sectors/industries) ────────
    sectors    = sorted({m["sector"]   for m in ticker_meta.values()
                         if m["sector"]   not in ("N/A", None, "")})
    industries = sorted({m["industry"] for m in ticker_meta.values()
                         if m["industry"] not in ("N/A", None, "")})
    build_settings_sheet(wb, sectors=sectors, industries=industries)

    # ── Build Dashboard ────────────────────────────────────────────
    build_dashboard(wb, tickers, ticker_meta)

    # ── Sheet order: Settings → Dashboard → SPY → rest ────────────
    ordered = ["Settings", "Dashboard"] + tickers
    wb._sheets.sort(key=lambda s: ordered.index(s.title)
                                  if s.title in ordered else len(ordered))

print(f"\nDone → {OUTPUT_FILE}")
